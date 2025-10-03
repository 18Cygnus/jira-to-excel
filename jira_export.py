import os
import math
import time
from typing import List, Dict, Any, Optional

import requests
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from gspread_dataframe import set_with_dataframe


from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

load_dotenv()

JIRA_BASE_URL = os.getenv("JIRA_BASE_URL", "").rstrip("/")
JIRA_EMAIL = os.getenv("JIRA_EMAIL")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN")
JIRA_FILTER_ID = os.getenv("JIRA_FILTER_ID")
JIRA_FILTER_NAME = os.getenv("JIRA_FILTER_NAME")
JIRA_JQL = os.getenv("JIRA_JQL")

PROJECT_ROOT = Path(__file__).resolve().parent
EXCEL_DIR = PROJECT_ROOT / "excels"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)  

# Tuning

MAX_RESULTS = 100
REQUEST_TIMEOUT = 30
RETRY_429_SECONDS = 10
OUTPUT_FILE = EXCEL_DIR / "jira_export.xlsx"
BULK_FETCH_CHUNK = 100  # Jira bulk fetch limit

SESSION = requests.Session()
SESSION.auth = (JIRA_EMAIL, JIRA_API_TOKEN)
SESSION.headers.update({"Accept": "application/json"})

def _require_env():
    missing = [k for k, v in {
        "JIRA_BASE_URL": JIRA_BASE_URL,
        "JIRA_EMAIL": JIRA_EMAIL,
        "JIRA_API_TOKEN": JIRA_API_TOKEN
    }.items() if not v]
    if missing:
        raise RuntimeError(f"Missing env vars: {', '.join(missing)}")
    
def _sleep_retry_after(resp):
    retry_after = resp.headers.get("Retry-After")
    try:
        wait = int(retry_after) if retry_after else RETRY_429_SECONDS
    except ValueError:
        wait = RETRY_429_SECONDS
    time.sleep(wait)

def search_issue_ids_by_jql(jql: str, max_results: int = MAX_RESULTS) -> List[str]:
    """
    Uses the new /rest/api/3/search/jql endpoint to collect issue IDs (or keys).
    Paginates with nextPageToken.
    """
    url = f"{JIRA_BASE_URL}/rest/api/3/search/jql"
    ids: List[str] = []
    next_token: Optional[str] = None

    while True:
        payload = {"jql": jql, "maxResults": max_results}
        if next_token:
            payload["nextPageToken"] = next_token

        r = SESSION.post(url, json=payload, timeout=REQUEST_TIMEOUT)
        if r.status_code == 429:
            _sleep_retry_after(r)
            continue
        if r.status_code != 200:
            raise RuntimeError(
                f"Failed to search issue IDs via /search/jql: {r.status_code} {r.text}"
            )
        
        data = r.json()
        issues = data.get("issues", [])
        # New API returns only ids/keys unless fields were requested
        for it in issues:
            ids.append(it.get("id") or it.get("key"))

        next_token = data.get("nextPageToken")
        if not next_token or not issues:
            break

    return ids

def bulk_fetch_issues(issue_ids_or_keys: List[str], fields: List[str]) -> List[Dict[str, Any]]:
    """
    Fetches full issue objects/fields via /rest/api/3/issue/bulkfetch in chunks
    """
    url = f"{JIRA_BASE_URL}/rest/api/3/issue/bulkfetch"
    out: List[Dict[str, Any]] = []

    for i in range(0, len(issue_ids_or_keys), BULK_FETCH_CHUNK):
        chunk = issue_ids_or_keys[i:i+BULK_FETCH_CHUNK]
        payload = {"issueIdsOrKeys": chunk, "fields": fields}

        while True:
            r = SESSION.post(url, json=payload, timeout=REQUEST_TIMEOUT)
            if r.status_code == 429:
                _sleep_retry_after(r)
                continue
            if r.status_code != 200:
                raise RuntimeError(
                    f"Bulk fetch failed (chunk {i}-{i+len(chunk)-1}): {r.status_code} {r.text}"
                )
            data = r.json()
            out.extend(data.get("issues", []))
            break

    return out

def fetch_all_issues(jql: str, fields: List[str]) -> List[Dict[str, Any]]:
    """Collect issue IDs for the given JQL and bulk fetch their fields."""
    ids = search_issue_ids_by_jql(jql, max_results=MAX_RESULTS)
    if not ids:
        return []
    return bulk_fetch_issues(ids, fields)

def dt_obj(dt_str):
    """
    Converts a datetime string from Jira to a datetime object.
    """
    if not dt_str:
        return None
    # contoh input: 2025-05-27T11:56:17.563+0700
    dt = datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%S.%f%z")
    # Buang timezone -> Excel-friendly (tetap jam lokal sesuai string)
    return dt.replace(tzinfo=None)

def d_obj(d_str):
    """
    Converts a date string from Jira to a date object.
    """
    return datetime.strptime(d_str, "%Y-%m-%d").date() if d_str else None

def flatten_issue(issue: Dict[str, Any]) -> Dict[str, Any]:
    """
    Converts Jira's nested issue JSON to flat columns
    add or remove fields as necessary
    """
    fields = issue.get("fields", {})
    def g(*keys, default=None):
        cur = fields
        for k in keys:
            cur = (cur or {}).get(k)
        return cur if cur is not None else default
    
    res_obj = fields.get("resolution")
    resolution_name = (
        res_obj.get("name") if isinstance(res_obj, dict) and res_obj.get("name") else "Unresolved"
    )
    
    return {
        "key": issue.get("key"),
        "work": g("summary"),
        "assignee": g("assignee", "displayName"),
        "reporter": g("reporter", "displayName"),
        "priority": g("priority", "name"),
        "status": g("status", "name"),
        "resolution": resolution_name,
        "created": dt_obj(g("created")),
        "updated": dt_obj(g("updated")),
        "due_date": d_obj(g("duedate"))
    }

def format_duration(seconds: float) -> str:
    total = int(round(seconds))
    m, s = divmod(total, 60)
    if m >= 60:
        h, m = divmod(m, 60)
        return f"{h} h {m} min {s} sec"
    return f"{m} min {s} sec"

def main():
    print(f"=== Jira Export Started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    _require_env()

    # determine source JQL
    if JIRA_JQL and JIRA_JQL.strip():
        active_jql = JIRA_JQL.strip()
        source_label = f"JQL '{active_jql}'"
        filter_id = None
    else:
        filter_id = None
        if JIRA_FILTER_ID:
            try:
                filter_id = int(JIRA_FILTER_ID)
            except ValueError:
                raise RuntimeError("JIRA_FILTER_ID must be a number")
        elif JIRA_FILTER_NAME:
            filter_id = find_filter_id_by_name(JIRA_FILTER_NAME)
            if filter_id is None:
                raise RuntimeError(f"Filter '{JIRA_FILTER_NAME}' not found")
        else:
            raise RuntimeError("Provide JIRA_JQL or JIRA_FILTER_ID or JIRA_FILTER_NAME in .env")
        active_jql = f"filter={filter_id}"
        if JIRA_FILTER_NAME:
            source_label = f"filter '{JIRA_FILTER_NAME}' (ID {filter_id})"
        else:
            source_label = f"filter ID {filter_id}"
    
    # choose the fields you want from Jira
    # include everything you use in flatten_issue()
    fields = [
        "summary", "assignee", "reporter", "priority", "status", "resolution",
        "created", "updated", "duedate", # change/remove if your instance differs
    ]

    print(f"Fetching issues for {source_label} ...")
    t0 = time.perf_counter() 
    issues = fetch_all_issues(active_jql, fields=fields)
    print(f"Fetched {len(issues)} issues")
    t_end = time.perf_counter()
    print(f"Total run time   : {format_duration(t_end - t0)}")

    rows = [flatten_issue(i) for i in issues]
    df = pd.DataFrame(rows)

    # sort & write to excel (sort by "created" column)
    if not df.empty:
        df.sort_values(by=["created"], ascending=False, inplace=True)
    
    # Update Google Sheets
    push_to_gsheet(df)

    with pd.ExcelWriter(str(OUTPUT_FILE), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Issues")
        
        # Auto-fit columns
        ws = writer.sheets["Issues"]
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # Add filtering dropdown for each column
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Tentukan range data (dari A1 sampai kolom terakhir & baris terakhir)
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName="IssuesTable", ref=ref)

        # # Configure column style
        # style = TableStyleInfo(
        #     name="TableStyleMedium9",
        #     showRowStripes=True,
        #     showColumnStripes=False
        # )
        # table.tableStyleInfo = style

        # Tambahkan table ke worksheet
        ws.add_table(table)

        # Freeze header row (opsional, supaya header tetap terlihat saat scroll)
        ws.freeze_panes = "A2"

        # === Format tanggal ===
        from openpyxl.styles import numbers

        # Cari index kolom (1-based untuk Excel)
        cols = {name: i+1 for i, name in enumerate(df.columns)}
        # created/updated: "Aug 15, 2025, 6:42 PM"
        fmt_dt = 'mmm d, yyyy, h:mm AM/PM'
        for row in range(2, ws.max_row + 1):
            if "created" in cols:
                ws.cell(row=row, column=cols["created"]).number_format = fmt_dt
            if "updated" in cols:
                ws.cell(row=row, column=cols["updated"]).number_format = fmt_dt
        # due_date: "Aug 15, 2025"
        if "due_date" in cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=cols["due_date"]).number_format = 'mmm d, yyyy'

    print(f"Exported {len(issues)} issues to {OUTPUT_FILE}")
    print(f"=== Jira Export Completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    print()  # Add empty line for separation between runs

def push_to_gsheet(df: pd.DataFrame):
    """Push DataFrame to Google Sheets (overwrite)"""
    gsheet_id = os.getenv("GSHEET_ID")
    gsheet_ws = os.getenv("GSHEET_WORKSHEET", "Sheet1")
    cred_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE")

    if not gsheet_id or not cred_file:
        print("Google sheets env not set; skipping gsheet update")
        return
    
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets"
        ]
        creds = Credentials.from_service_account_file(cred_file, scopes=scopes)
        gc = gspread.authorize(creds)

        # Open the spreadsheet by ID
        sh = gc.open_by_key(gsheet_id)
        print(f"Successfully opened Google Sheet: {sh.title}")

        # Get the first worksheet (index 0) regardless of name
        ws = sh.get_worksheet(0)
        print(f"Using worksheet: {ws.title}")

        # Clear existing content and resize sheet if needed
        ws.clear()
        
        # Calculate required dimensions
        required_rows = len(df) + 1  # +1 for header
        required_cols = len(df.columns)
        
        # Resize sheet if necessary
        current_rows = ws.row_count
        current_cols = ws.col_count
        
        if required_rows > current_rows or required_cols > current_cols:
            new_rows = max(required_rows, current_rows)
            new_cols = max(required_cols, current_cols)
            print(f"Resizing sheet from {current_rows}x{current_cols} to {new_rows}x{new_cols}")
            ws.resize(rows=new_rows, cols=new_cols)

        # Convert all non-JSON serializable types to string with better formatting
        df_copy = df.copy()
        for col in df_copy.columns:
            if col in ['created', 'updated']:
                # Format datetime columns like "Sep 08, 2025, 2:55 PM"
                df_copy[col] = df_copy[col].dt.strftime('%b %d, %Y, %I:%M %p')
            elif col == 'due_date':
                # Format date column like "Sep 08, 2025" 
                df_copy[col] = df_copy[col].apply(
                    lambda x: x.strftime('%b %d, %Y') if pd.notnull(x) else ''
                )
            elif df_copy[col].dtype == 'datetime64[ns]':
                # Other datetime columns with standard format
                df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            elif 'datetime' in str(df_copy[col].dtype):
                df_copy[col] = df_copy[col].astype(str)
            elif df_copy[col].dtype == 'object':
                # Handle date objects and other non-serializable objects
                df_copy[col] = df_copy[col].astype(str)

        # Replace NaN/None values with empty strings
        df_copy = df_copy.fillna('')

        # Push DataFrame to Google Sheets in smaller chunks to avoid timeouts
        print(f"Pushing {len(df_copy)} rows to Google Sheet...")
        
        # First, set headers (using new parameter order)
        headers = [list(df_copy.columns)]
        ws.update(values=headers, range_name='A1')
        
        # Then push data in chunks of 1000 rows
        chunk_size = 1000
        for i in range(0, len(df_copy), chunk_size):
            chunk = df_copy.iloc[i:i+chunk_size]
            start_row = i + 2  # +2 because row 1 is header and sheets are 1-indexed
            
            # Convert chunk to list of lists and ensure all values are JSON serializable
            data = []
            for _, row in chunk.iterrows():
                row_data = []
                for value in row:
                    if pd.isna(value) or value is None:
                        row_data.append('')
                    else:
                        row_data.append(str(value))
                data.append(row_data)
            
            # Update the range (using new parameter order)
            end_row = start_row + len(data) - 1
            end_col_letter = chr(ord('A') + len(df_copy.columns) - 1)  # Calculate column letter
            range_name = f'A{start_row}:{end_col_letter}{end_row}'
            
            ws.update(values=data, range_name=range_name)
            print(f"Uploaded rows {start_row}-{end_row}")

        # Apply formatting to make the spreadsheet look more professional
        print("Applying formatting to the spreadsheet...")
        
        # Format header row (bold, background color)
        header_range = f'A1:{chr(ord("A") + len(df_copy.columns) - 1)}1'
        ws.format(header_range, {
            "backgroundColor": {"red": 0.2, "green": 0.2, "blue": 0.2},
            "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True}
        })
        
        # Auto-resize columns
        # ws.columns_auto_resize(0, len(df_copy.columns) - 1)
        
        # Freeze header row
        ws.freeze(rows=1)
        
        print("Formatting applied successfully!")

        print(f"Successfully pushed {len(df)} rows to Google Sheet '{ws.title}'.")
        print(f"Sheet URL: https://docs.google.com/spreadsheets/d/{sh.id}/edit")
    
    except Exception as e:
        print(f"Failed to push to Google Sheets: {e}")
        print("Continuing with Excel export only...")

if __name__ == "__main__":
    main()